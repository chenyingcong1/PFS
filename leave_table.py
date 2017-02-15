
from openpyxl import load_workbook
from os import path
from datetime import datetime

otsheetDemo = path.realpath('./static/sheet/otsheetDemo.xlsx')
hotsheetDemo = path.realpath('./static/sheet/hotsheetDemo.xlsx')
annsheetDemo = path.realpath('./static/sheet/annsheetDemo.xlsx')
bsnssheetDemo = path.realpath('./static/sheet/bsnssheetDemo.xlsx')
def make_otsheet(names=(),
                 department=(),
                 position=(),
                 thing=(),
                 date=(),
                 place=(),
                 ottime=(),
                 datetimeStart=(),
                 datetimeEnd=()):

    wb = load_workbook(otsheetDemo)
    ws = wb.get_sheet_by_name('日常加班单')
    ws["B2"] = ws["B10"] = names
    ws["D2"] = ws["D10"] = department
    ws["F2"] = ws["F10"] = position
    ws["B3"] = ws["B11"] = thing
    ws["B4"] = ws["B12"] = date
    ws["D4"] = ws["D12"] = place
    ws["B5"] = ws["B13"] = ottime+"小时"
    ws["D5"] = ws["D13"] = "{}至{},共计{}小时".format(datetimeStart,datetimeEnd,ottime)
    wb.save('./static/sheet/otsheet.xlsx')

def make_hotsheet(name1=(),name2=(),name3=(),name4=(),name5=(),
                  position1=(),position2=(),position3=(),position4=(),position5=(),
                  datetimeStart1=(),datetimeStart2=(),datetimeStart3=(),datetimeStart4=(),datetimeStart5=(),
                  datetimeEnd1=(),datetimeEnd2=(),datetimeEnd3=(),datetimeEnd4=(),datetimeEnd5=(),
                  thing=()):
    wb = load_workbook(hotsheetDemo)
    ws = wb.get_sheet_by_name('假日加班单')
    ws["B3"] = ws["B14"] = name1
    ws["B4"] = ws["B15"] = name2
    ws["B5"] = ws["B16"] = name3
    ws["B6"] = ws["B17"] = name4
    ws["B7"] = ws["B18"] = name5
    ws["C3"] = ws["C14"] = position1
    ws["C4"] = ws["C15"] = position2
    ws["C5"] = ws["C16"] = position3
    ws["C6"] = ws["C17"] = position4
    ws["C7"] = ws["C18"] = position5
    ws["D3"] = ws["D14"] = "{}--{}".format(datetimeStart1,datetimeEnd1)
    ws["D4"] = ws["D15"] = "{}--{}".format(datetimeStart2, datetimeEnd2)
    ws["D5"] = ws["D16"] = "{}--{}".format(datetimeStart3, datetimeEnd3)
    ws["D6"] = ws["D17"] = "{}--{}".format(datetimeStart4, datetimeEnd4)
    ws["D7"] = ws["D18"] = "{}--{}".format(datetimeStart5, datetimeEnd5)
    ws["A8"] = ws["A19"] = "加班事由及地点：{}".format(thing)
    wb.save('./static/sheet/hotsheet.xlsx')

def make_annsheet(names=(),
                  department=(),
                  position=(),
                  thing=(),
                  datetimeStart=(),
                  datetimeEnd=(),
                  thing1=()):
    wb = load_workbook(annsheetDemo)
    ws = wb.get_sheet_by_name('请休假申请表')
    ws["B2"] = ws["B14"] = names
    ws["D2"] = ws["D14"] = department
    ws["F2"] = ws["F14"] = position
    ws["B4"] = ws["B16"] = thing
    day = datetime.strptime(datetimeEnd,'%Y-%m-%d %H:%M')-datetime.strptime(datetimeStart,'%Y-%m-%d %H:%M')
    ws["B5"] = ws["B17"] = "由{}至{}，共计{}天".format(datetimeStart,datetimeEnd,day.days)
    ws["A6"] = ws["A18"] = "工作交接安排：{}".format(thing1)
    wb.save('./static/sheet/annsheet.xlsx')

def make_bsnssheet(names=(),
                   department=(),
                   personnel=(),
                   datetimeStart=(),
                   datetimeEnd=(),
                   thing=(),
                   thing1=(),
                   inform=(),
                   outform=(),
                   announcement=(),
                   thing2=(),
                   thing3=()):
    wb = load_workbook(bsnssheetDemo)
    ws = wb.get_sheet_by_name('业务调整审批登记表')
    ws["C2"] = names
    ws["F2"] = department
    ws["H2"] = personnel
    ws["C3"] = "{}至{}".format(datetimeStart,datetimeEnd)
    ws["B5"] = thing
    ws["B6"] = thing1
    ws["B7"] = inform
    ws["E7"] = outform
    ws["I7"] = announcement
    ws["C8"] = thing2
    ws["C9"] = thing3
    wb.save('./static/sheet/bsnssheet.xlsx')



