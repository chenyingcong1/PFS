
from openpyxl import load_workbook
from os import path

Demo = path.realpath('./static/sheet/hotsheetDemo.xlsx')


wb = load_workbook(Demo)
ws = wb.get_sheet_by_name('假日加班单')
ws["B3"] = ws["B14"] = "测试"

wb.save('./static/sheet/hotsheet.xlsx')